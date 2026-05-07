from pathlib import Path
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

IMAGE_FILE = Path("image/sample.png")
EXCEL_FILE = Path("sample.xlsx")
SHEET_NAME = "Sheet1"
CELL_RANGE = "B11"


def insert_image_to_excel(
    excel_path: Path,
    image_path: Path,
    sheet_name: str,
    cell: str,
) -> None:
    """
    Excelの指定シート・セルに画像を挿入して上書き保存する。

    シートが存在しない場合は新規作成する。
    """
    wb = load_workbook(excel_path)

    # シートがなければ作成、あればそのまま取得
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(title=sheet_name)

    ws.add_image(Image(image_path), cell)
    wb.save(excel_path)
    print(f"{sheet_name} シートの {cell} セルに画像を追加")


insert_image_to_excel(EXCEL_FILE, IMAGE_FILE, SHEET_NAME, CELL_RANGE)
